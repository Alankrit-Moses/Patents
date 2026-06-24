"""Shared CAGR extraction utilities for Pattern 2."""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def find_repository_root(anchor: Path) -> Path:
    start = anchor.resolve().parent
    for candidate in (start, *start.parents):
        if (candidate / "Patterns").is_dir() and (candidate / "tables").is_dir():
            return candidate
    raise FileNotFoundError("Could not locate the workspace root.")


def normalize_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.year
    try:
        return int(str(value).strip()[:4])
    except (TypeError, ValueError):
        return None


def extract_cagr_comparison(
    *,
    workbook_path: Path,
    start_year: int,
    end_year: int,
    sheet_specs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Calculate CAGR from distinct family counts for each configured entity."""
    if end_year <= start_year:
        raise ValueError("end_year must be greater than start_year.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    output: list[dict[str, Any]] = []
    try:
        for sheet_spec in sheet_specs:
            sheet_name = sheet_spec["sheet_name"]
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"Worksheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
            header_row = sheet_spec["header_row"]
            header_values = next(
                sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
            )
            headers = {
                str(value).strip(): index
                for index, value in enumerate(header_values)
                if value is not None
            }
            family_column = sheet_spec["family_column"]
            year_column = sheet_spec["year_column"]
            entities = sheet_spec["entities"]
            required = {family_column, year_column}
            for entity in entities:
                required.update(entity.get("filters", {}).keys())
            missing = sorted(required - headers.keys())
            if missing:
                raise KeyError(f"Missing columns in {sheet_name}: {', '.join(missing)}")

            families: dict[tuple[str, int], set[Any]] = defaultdict(set)
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                year = normalize_year(row[headers[year_column]])
                if year not in {start_year, end_year}:
                    continue
                family_id = row[headers[family_column]]
                if family_id is None:
                    continue
                for entity in entities:
                    filters = entity.get("filters", {})
                    if all(row[headers[column]] == expected for column, expected in filters.items()):
                        families[(entity["entity"], year)].add(family_id)

            periods = end_year - start_year
            for entity in entities:
                label = entity["entity"]
                start_count = len(families[(label, start_year)])
                end_count = len(families[(label, end_year)])
                if start_count <= 0:
                    raise ValueError(
                        f"CAGR is undefined for {label}: start count in {start_year} is {start_count}."
                    )
                cagr_percent = ((end_count / start_count) ** (1 / periods) - 1) * 100
                output.append(
                    {
                        "entity": label,
                        "start_year": start_year,
                        "end_year": end_year,
                        "cagr_percent": round(cagr_percent, 1),
                    }
                )
        return output
    finally:
        workbook.close()


def emit_result(result: dict[str, Any], output_path: Path | None) -> None:
    if output_path is None:
        print(json.dumps(result, indent=2, ensure_ascii=False))
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        rows = result["tabular_example"]
        if not rows:
            raise ValueError("Cannot write an empty comparison table.")
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)
    else:
        output_path.write_text(
            json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
    print(f"Saved {output_path}")
