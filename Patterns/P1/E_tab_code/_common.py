"""Shared utilities for extracting P1 tabular examples from source workbooks."""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def find_repository_root(anchor: Path) -> Path:
    """Find the workspace root without depending on the current directory."""
    start = anchor.resolve().parent
    for candidate in (start, *start.parents):
        if (candidate / "Patterns").is_dir() and (candidate / "tables").is_dir():
            return candidate
    raise FileNotFoundError("Could not locate a repository root containing Patterns/ and tables/.")


def normalize_year(value: Any) -> int | None:
    """Convert Excel dates, numeric years, or ISO-like strings into a four-digit year."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.year
    if isinstance(value, (int, float)) and 1000 <= int(value) <= 9999:
        return int(value)
    match = re.match(r"\s*(\d{4})", str(value))
    return int(match.group(1)) if match else None


def extract_distinct_family_counts(
    *,
    workbook_path: Path,
    sheet_name: str,
    header_row: int,
    family_column: str,
    year_column: str,
    start_year: int,
    end_year: int,
    filters: dict[str, Any] | None = None,
    count_field: str = "distinct_patent_family_count",
    include_family_ids: bool = False,
) -> list[dict[str, Any]]:
    """Return COUNT(DISTINCT family ID) by year for rows satisfying exact filters."""
    filters = filters or {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_values)
            if value is not None
        }

        required = {family_column, year_column, *filters.keys()}
        missing = sorted(required - headers.keys())
        if missing:
            raise KeyError(f"Missing required column(s) in {sheet_name}: {', '.join(missing)}")

        families_by_year: dict[int, set[Any]] = defaultdict(set)
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if any(row[headers[column]] != expected for column, expected in filters.items()):
                continue
            family_id = row[headers[family_column]]
            year = normalize_year(row[headers[year_column]])
            if family_id is None or year is None or not start_year <= year <= end_year:
                continue
            families_by_year[year].add(family_id)

        result: list[dict[str, Any]] = []
        for year in range(start_year, end_year + 1):
            family_ids = families_by_year.get(year, set())
            record: dict[str, Any] = {
                "year": year,
                count_field: len(family_ids),
            }
            if include_family_ids:
                record["patent_family_ids"] = sorted(family_ids, key=str)
            result.append(record)
        return result
    finally:
        workbook.close()


def emit_json(result: dict[str, Any], output_path: Path | None) -> None:
    """Print JSON by default; save CSV or JSON when --output is supplied."""
    payload = json.dumps(result, indent=2, ensure_ascii=False)
    if output_path is None:
        print(payload)
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        rows = result["tabular_example"]
        if not rows:
            raise ValueError("Cannot write a CSV for an empty tabular example.")
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)
        print(f"Saved {output_path}")
        return
    output_path.write_text(payload + "\n", encoding="utf-8")
    print(f"Saved {output_path}")
